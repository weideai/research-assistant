import hashlib
import io
import json
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app import db
from app.export_service import (
    build_archive_export,
    build_docx_export,
    build_json_export,
    build_markdown_export,
    build_xlsx_export,
    experiment_payload,
)
from app.models import (
    BatchStep,
    Experiment,
    ExperimentBatch,
    ExperimentStep,
    ResearchProject,
    User,
)
from app.project_package import (
    PACKAGE_FORMAT,
    PACKAGE_SCHEMA_VERSION,
    build_project_package,
    import_project_package,
)


def _seed_execution_steps(auth, app):
    auth.register()
    with app.app_context():
        user = User.query.one()
        project = ResearchProject(user_id=user.id, title="执行步骤导出项目")
        db.session.add(project)
        db.session.flush()
        experiment = Experiment(
            user_id=user.id,
            project_id=project.id,
            title="执行步骤导出实验",
            code="EXP-STEP-EXPORT",
            objective="验证计划定义与执行状态分离。",
        )
        db.session.add(experiment)
        db.session.flush()
        plan_step = ExperimentStep(
            experiment_id=experiment.id,
            position=1,
            title="计划步骤定义",
            description="计划层只保存定义",
            operator="计划负责人",
            planned_date=date(2026, 7, 20),
        )
        db.session.add(plan_step)
        db.session.flush()
        batch_a = ExperimentBatch(experiment_id=experiment.id, batch_code="RUN-A")
        batch_b = ExperimentBatch(experiment_id=experiment.id, batch_code="RUN-B")
        db.session.add_all((batch_a, batch_b))
        db.session.flush()
        step_a = BatchStep.from_plan_step(batch_a.id, plan_step)
        step_a.title = "执行快照 A"
        step_a.is_done = True
        step_a.completed_date = date(2026, 7, 21)
        step_b = BatchStep.from_plan_step(batch_b.id, plan_step)
        step_b.title = "执行快照 B"
        db.session.add_all((step_a, step_b))
        db.session.commit()
        return user.id, project.id, experiment.id


def test_json_and_markdown_export_separate_plan_definitions_from_execution_status(auth, app):
    _user_id, _project_id, experiment_id = _seed_execution_steps(auth, app)
    with app.app_context():
        experiment = db.session.get(Experiment, experiment_id)
        payload = experiment_payload(experiment, exported_at=datetime(2026, 7, 24, 12, 0))

        assert payload["schema_version"] == 3
        assert set(payload["steps"][0]) == {
            "id", "position", "title", "description", "operator", "planned_date",
        }
        executions = {batch["batch_code"]: batch for batch in payload["batches"]}
        assert executions["RUN-A"]["steps"][0]["title"] == "执行快照 A"
        assert executions["RUN-A"]["steps"][0]["is_done"] is True
        assert executions["RUN-A"]["steps"][0]["completed_date"] == "2026-07-21"
        assert executions["RUN-B"]["steps"][0]["title"] == "执行快照 B"
        assert executions["RUN-B"]["steps"][0]["is_done"] is False

        json_payload = json.loads(build_json_export(experiment))
        assert "is_done" not in json_payload["steps"][0]
        assert json_payload["batches"][0]["steps"]

        markdown = build_markdown_export(experiment)
        assert "## 计划步骤定义" in markdown
        assert "#### 执行步骤" in markdown
        assert "执行快照 A" in markdown
        assert "已完成" in markdown
        assert "执行快照 B" in markdown
        assert "待完成" in markdown


def test_word_excel_and_zip_exports_include_execution_scoped_steps(auth, app):
    _user_id, _project_id, experiment_id = _seed_execution_steps(auth, app)
    with app.app_context():
        experiment = db.session.get(Experiment, experiment_id)

        with zipfile.ZipFile(io.BytesIO(build_docx_export(experiment))) as document:
            document_xml = document.read("word/document.xml").decode("utf-8")
        assert "计划步骤定义" in document_xml
        assert "执行步骤" in document_xml
        assert "执行快照 A" in document_xml
        assert "已完成" in document_xml

        workbook = load_workbook(io.BytesIO(build_xlsx_export(experiment)), read_only=True, data_only=True)
        try:
            assert "执行步骤" in workbook.sheetnames
            plan_rows = list(workbook["实验步骤"].iter_rows(values_only=True))
            execution_rows = list(workbook["执行步骤"].iter_rows(values_only=True))
        finally:
            workbook.close()
        assert plan_rows[0] == ("序号", "步骤", "计划执行人", "计划日期", "说明")
        assert "完成日期" not in plan_rows[0]
        assert {row[1] for row in execution_rows[1:]} == {"RUN-A", "RUN-B"}
        assert {row[6] for row in execution_rows[1:]} == {"执行快照 A", "执行快照 B"}

        archive_file = build_archive_export(experiment, lambda _attachment: None)
        try:
            with zipfile.ZipFile(archive_file) as archive:
                archive_payload = json.loads(archive.read("experiment.json"))
                report = archive.read("report.md").decode("utf-8-sig")
        finally:
            archive_file.close()
        assert archive_payload["schema_version"] == 3
        assert archive_payload["batches"][0]["steps"]
        assert "执行快照 A" in report
        assert "执行快照 B" in report


def test_project_package_round_trip_preserves_execution_steps(auth, app):
    user_id, project_id, _experiment_id = _seed_execution_steps(auth, app)
    package_path = None
    try:
        with app.app_context():
            project = db.session.get(ResearchProject, project_id)
            package_path, _manifest = build_project_package(
                project, app.config["ATTACHMENT_UPLOAD_DIR"]
            )
            with zipfile.ZipFile(package_path) as archive:
                project_payload = json.loads(archive.read("project.json"))
            packaged_experiment = project_payload["experiments"][0]
            assert "is_done" not in packaged_experiment["steps"][0]
            assert "completed_date" not in packaged_experiment["steps"][0]
            packaged_batches = {
                batch["batch_code"]: batch for batch in packaged_experiment["batches"]
            }
            assert packaged_batches["RUN-A"]["steps"][0]["is_done"] is True
            assert packaged_batches["RUN-A"]["steps"][0]["completed_date"] == "2026-07-21"
            assert packaged_batches["RUN-B"]["steps"][0]["is_done"] is False

            with Path(package_path).open("rb") as package_file:
                imported_project, _manifest = import_project_package(
                    package_file, user_id, app.config["ATTACHMENT_UPLOAD_DIR"]
                )
            imported_experiment = Experiment.query.filter_by(project_id=imported_project.id).one()
            imported_batches = {batch.batch_code: batch for batch in imported_experiment.batches}
            assert imported_batches["RUN-A"].steps[0].title == "执行快照 A"
            assert imported_batches["RUN-A"].steps[0].is_done is True
            assert imported_batches["RUN-A"].steps[0].source_step_id == imported_experiment.steps[0].id
            assert imported_batches["RUN-B"].steps[0].title == "执行快照 B"
            assert imported_batches["RUN-B"].steps[0].is_done is False
    finally:
        if package_path:
            Path(package_path).unlink(missing_ok=True)


def test_legacy_project_package_copies_plan_completion_to_every_execution(auth, app):
    auth.register()
    with app.app_context():
        user_id = User.query.one().id
    project_payload = {
        "project": {"title": "旧项目包"},
        "experiments": [{
            "title": "旧实验",
            "steps": [{
                "position": 1,
                "title": "旧版完成步骤",
                "description": "旧格式将状态保存在计划步骤上",
                "operator": "旧负责人",
                "planned_date": "2026-07-20",
                "completed_date": "2026-07-21",
                "is_done": True,
            }],
            "batches": [
                {"batch_code": "LEGACY-A"},
                {"batch_code": "LEGACY-B"},
            ],
        }],
    }
    project_json = json.dumps(project_payload, ensure_ascii=False).encode("utf-8")
    manifest = {
        "format": PACKAGE_FORMAT,
        "schema_version": 1,
        "project_title": "旧项目包",
        "created_at": "2026-07-24T12:00:00Z",
        "entries": {"project.json": hashlib.sha256(project_json).hexdigest()},
    }
    package_file = io.BytesIO()
    with zipfile.ZipFile(package_file, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        archive.writestr("project.json", project_json)
    package_file.seek(0)

    with app.app_context():
        imported_project, _manifest = import_project_package(
            package_file, user_id, app.config["ATTACHMENT_UPLOAD_DIR"]
        )
        imported_experiment = Experiment.query.filter_by(project_id=imported_project.id).one()
        assert len(imported_experiment.steps) == 1
        assert not hasattr(imported_experiment.steps[0], "is_done")
        assert {batch.batch_code for batch in imported_experiment.batches} == {
            "LEGACY-A", "LEGACY-B",
        }
        for batch in imported_experiment.batches:
            assert len(batch.steps) == 1
            assert batch.steps[0].source_step_id == imported_experiment.steps[0].id
            assert batch.steps[0].is_done is True
            assert batch.steps[0].completed_date == date(2026, 7, 21)
