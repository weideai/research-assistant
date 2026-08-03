import io
import json
import re
import zipfile
from datetime import date

from openpyxl import load_workbook

from app import db
from app.models import (
    BatchStep,
    Experiment,
    ExperimentAttachment,
    ExperimentBatch,
    ExperimentParameter,
    ExperimentRecord,
    RecordParameter,
    ResearchProject,
)


ACTIVE_ATTACHMENT = "raw/result-a.csv"
DELETED_ATTACHMENT = "discarded/deleted-attachment.txt"
DELETED_RECORD_ATTACHMENT = "discarded/deleted-record.csv"
DELETED_RECORD_CONTENT = "SOFT-DELETED-RECORD"
UNASSIGNED_RECORD_CONTENT = "ORPHAN-LEGACY-RECORD"


def _post_record(client, experiment_id, batch_id, *, record_date, content, result, filename, file_data):
    response = client.post(f"/experiments/{experiment_id}/records", data={
        "batch_id": str(batch_id),
        "record_date": record_date,
        "operator": "研究员",
        "conditions": "37°C，5% CO2",
        "content": content,
        "result": result,
        "remark": "建议增加一次独立重复。",
        "record_parameter_name": ["曝光时间"],
        "record_parameter_value": ["30"],
        "record_parameter_unit": ["s"],
        "record_parameter_notes": ["未饱和"],
        "attachment_category": "原始数据",
        "files": (io.BytesIO(file_data), filename),
    }, content_type="multipart/form-data")
    assert response.status_code == 302


def _create_complete_experiment(client, auth, app):
    auth.register()
    client.post("/experiments", data={
        "title": "药物处理后 WB 验证",
        "code": "EXP-EXPORT-01",
        "owner": "研究员",
        "status": "进行中",
        "start_date": "2026-07-20",
        "end_date": "2026-07-24",
        "objective": "验证目标蛋白表达变化。",
    })
    with app.app_context():
        experiment = Experiment.query.one()
        experiment_id = experiment.id
        db.session.add(ExperimentParameter(
            experiment_id=experiment_id,
            position=1,
            name="药物浓度",
            value="5",
            unit="μM",
            notes="终浓度",
        ))
        db.session.commit()

    for batch_data in (
        {
            "batch_code": "BATCH-A",
            "repeat_kind": "生物学重复",
            "repeat_number": "1",
            "group_name": "对照组",
            "operator": "研究员 A",
            "start_date": "2026-07-20",
        },
        {
            "batch_code": "BATCH-B",
            "repeat_kind": "生物学重复",
            "repeat_number": "2",
            "group_name": "处理组",
            "operator": "研究员 B",
            "start_date": "2026-07-22",
        },
    ):
        response = client.post(f"/experiments/{experiment_id}/batches", data=batch_data)
        assert response.status_code == 302

    with app.app_context():
        batch_ids = {
            batch.batch_code: batch.id
            for batch in ExperimentBatch.query.filter_by(experiment_id=experiment_id).all()
        }

    response = client.post(f"/experiments/{experiment_id}/steps", data={
        "title": "药物处理",
        "operator": "研究员",
        "planned_date": "2026-07-20",
        "description": "处理 24h",
    })
    assert response.status_code == 302

    _post_record(
        client,
        experiment_id,
        batch_ids["BATCH-A"],
        record_date="2026-07-21",
        content="完成第一批处理并采集原始数据。",
        result="BATCH-A-RESULT",
        filename=ACTIVE_ATTACHMENT,
        file_data=b"sample,value\nA,1\n",
    )
    _post_record(
        client,
        experiment_id,
        batch_ids["BATCH-B"],
        record_date="2026-07-23",
        content="完成第二批处理并采集图像。",
        result="BATCH-B-RESULT",
        filename="images/result-b.png",
        file_data=b"not-a-preview-but-valid-test-data",
    )
    _post_record(
        client,
        experiment_id,
        batch_ids["BATCH-B"],
        record_date="2026-07-24",
        content=DELETED_RECORD_CONTENT,
        result="DELETED-RESULT",
        filename=DELETED_RECORD_ATTACHMENT,
        file_data=b"deleted,record\n",
    )

    with app.app_context():
        first_record = ExperimentRecord.query.filter_by(result="BATCH-A-RESULT").one()
        first_record_id = first_record.id
    response = client.post(f"/records/{first_record_id}/attachments", data={
        "attachment_category": "实验文档",
        "attachment_folder": "discarded",
        "files": (io.BytesIO(b"deleted attachment"), "deleted-attachment.txt"),
    }, content_type="multipart/form-data")
    assert response.status_code == 302

    with app.app_context():
        deleted_record = ExperimentRecord.query.filter_by(content=DELETED_RECORD_CONTENT).one()
        deleted_record.is_deleted = True
        deleted_attachment = ExperimentAttachment.query.filter_by(
            record_id=first_record_id,
            relative_path=DELETED_ATTACHMENT,
        ).one()
        deleted_attachment.is_deleted = True
        experiment = db.session.get(Experiment, experiment_id)
        wrong_experiment = Experiment(
            user_id=experiment.user_id,
            project_id=experiment.project_id,
            title="错误执行归属",
        )
        db.session.add(wrong_experiment)
        db.session.flush()
        wrong_batch = ExperimentBatch(
            experiment_id=wrong_experiment.id,
            batch_code="WRONG-OWNER",
        )
        db.session.add(wrong_batch)
        db.session.flush()
        db.session.add(ExperimentRecord(
            experiment_id=experiment_id,
            batch_id=wrong_batch.id,
            record_date=date(2026, 7, 19),
            operator="历史研究员",
            conditions="旧系统导入",
            content=UNASSIGNED_RECORD_CONTENT,
            result="待归档",
            remark="迁移后应归入历史执行。",
        ))
        db.session.commit()
        assert ExperimentAttachment.query.count() == 4
        assert RecordParameter.query.count() == 3
    return experiment_id


def _docx_xml(data):
    with zipfile.ZipFile(io.BytesIO(data)) as document:
        assert "word/document.xml" in document.namelist()
        return document.read("word/document.xml").decode("utf-8")


def _workbook_rows(data):
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return {
            sheet.title: list(sheet.iter_rows(values_only=True))
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _workbook_text(data):
    rows = _workbook_rows(data)
    return "\n".join(
        str(value)
        for sheet_rows in rows.values()
        for row in sheet_rows
        for value in row
        if value is not None
    )


def test_experiment_export_picker_lists_supported_formats(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    response = client.get(f"/experiments/{experiment_id}")
    assert response.status_code == 200
    for label in ("Markdown 报告", "PDF 实验记录", "Word 文档", "Excel 工作簿", "JSON 结构化数据", "ZIP 完整归档", "科研档案"):
        assert label.encode() in response.data


def _docx_parts(data):
    with zipfile.ZipFile(io.BytesIO(data)) as document:
        return {
            name: document.read(name).decode("utf-8", errors="ignore")
            for name in document.namelist()
            if name.endswith(".xml")
        }


def _without_export_timestamp(markup):
    """Exports embed their own export time, so mask it before comparing renders."""
    return re.sub(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}", "<EXPORTED-AT>", markup)


REPORT_TEMPLATE_FINGERPRINTS = {
    "research": ("RESEARCH RECORD", "2166F3"),
    "notebook": ("LAB NOTEBOOK", "167C80"),
    "compact": ("RESULT SUMMARY", "A15C00"),
}


def test_word_export_honours_the_selected_report_template(client, auth, app):
    """The template picker must change Word output, not only PDF output."""
    experiment_id = _create_complete_experiment(client, auth, app)

    rendered = {}
    for template_key, (kicker, accent) in REPORT_TEMPLATE_FINGERPRINTS.items():
        response = client.get(
            f"/experiments/{experiment_id}/export?format=docx&report_template={template_key}"
        )
        assert response.status_code == 200
        parts = _docx_parts(response.data)
        header = "".join(markup for name, markup in parts.items() if "header" in name)
        assert kicker in header, f"{template_key}: 页眉未使用该模板的 kicker"
        assert f'w:fill="{accent}"' in parts["word/document.xml"], (
            f"{template_key}: 表头未使用该模板的强调色"
        )
        rendered[template_key] = _without_export_timestamp(parts["word/document.xml"])

    assert rendered["research"] != rendered["notebook"]
    assert rendered["notebook"] != rendered["compact"]

    unknown = client.get(
        f"/experiments/{experiment_id}/export?format=docx&report_template=does-not-exist"
    )
    assert unknown.status_code == 200
    unknown_xml = _without_export_timestamp(_docx_parts(unknown.data)["word/document.xml"])
    assert unknown_xml == rendered["research"]


def test_record_word_export_honours_the_selected_report_template(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    with app.app_context():
        record_id = ExperimentRecord.query.filter_by(
            experiment_id=experiment_id, is_deleted=False
        ).first().id

    rendered = {}
    for template_key, (kicker, accent) in REPORT_TEMPLATE_FINGERPRINTS.items():
        response = client.get(
            f"/records/{record_id}/export?format=docx&report_template={template_key}"
        )
        assert response.status_code == 200
        parts = _docx_parts(response.data)
        header = "".join(markup for name, markup in parts.items() if "header" in name)
        assert kicker in header
        assert f'w:fill="{accent}"' in parts["word/document.xml"]
        rendered[template_key] = _without_export_timestamp(parts["word/document.xml"])

    assert len(set(rendered.values())) == len(REPORT_TEMPLATE_FINGERPRINTS)


def test_record_report_page_word_and_pdf_include_batch_execution_steps(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    with app.app_context():
        batch = ExperimentBatch.query.filter_by(
            experiment_id=experiment_id, batch_code="BATCH-A",
        ).one()
        db.session.add(BatchStep(
            batch_id=batch.id,
            position=1,
            title="本批次裂解并定量",
            description="使用本批次校准后的裂解条件。",
            operator="执行研究员",
            planned_date=date(2026, 7, 21),
            completed_date=date(2026, 7, 21),
            is_done=True,
        ))
        record = ExperimentRecord.query.filter_by(result="BATCH-A-RESULT").one()
        record_id = record.id
        db.session.commit()

    page = client.get(f"/experiments/{experiment_id}/reports?record_id={record_id}")
    assert page.status_code == 200
    assert "本批次实验步骤".encode() in page.data
    assert "本批次裂解并定量".encode() in page.data
    assert "实际执行记录".encode() in page.data
    assert "独立执行快照".encode() not in page.data

    word = client.get(f"/records/{record_id}/export?format=docx")
    assert word.status_code == 200
    document_xml = _docx_xml(word.data)
    assert "本批次实验步骤" in document_xml
    assert "本批次裂解并定量" in document_xml
    assert "已完成" in document_xml

    pdf = client.get(f"/records/{record_id}/export?format=pdf")
    assert pdf.status_code == 200
    assert pdf.data.startswith(b"%PDF")


def test_selected_project_reports_are_previewed_before_export(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    with app.app_context():
        experiment = db.session.get(Experiment, experiment_id)
        project = ResearchProject(
            user_id=experiment.user_id,
            title="细胞应答项目",
            code="PROJ-REPORT",
        )
        db.session.add(project)
        db.session.flush()
        experiment.project_id = project.id
        project_id = project.id
        batch = ExperimentBatch.query.filter_by(
            experiment_id=experiment_id, batch_code="BATCH-A",
        ).one()
        db.session.add(BatchStep(
            batch_id=batch.id,
            position=1,
            title="预览中的裂解步骤",
            description="核对预览和导出内容一致。",
            operator="执行研究员",
            planned_date=date(2026, 7, 21),
            completed_date=date(2026, 7, 21),
            is_done=True,
        ))
        selected_record = ExperimentRecord.query.filter_by(result="BATCH-A-RESULT").one()
        unselected_record = ExperimentRecord.query.filter_by(result="BATCH-B-RESULT").one()
        selected_record_id = selected_record.id
        unselected_record_id = unselected_record.id
        db.session.commit()

    index = client.get(f"/experiment-reports?project_id={project_id}")
    assert index.status_code == 200
    assert b'action="/experiment-reports/export-preview"' in index.data
    assert b"data-bulk-select-all-matches" in index.data
    assert "选择当前筛选全部".encode() in index.data
    assert "预览所选报告".encode() in index.data
    assert f'/projects/{project_id}/reports/export'.encode() not in index.data
    assert index.data.count(b'name="record_ids"') == 3

    preview = client.post("/experiment-reports/export-preview", data={
        "selection_scope": "page",
        "record_ids": [str(selected_record_id)],
        "project_id": str(project_id),
        "q": "",
        "page": "1",
        "per_page": "10",
    })
    assert preview.status_code == 200
    assert "批量报告预览".encode() in preview.data
    assert "已载入 1 份实验报告".encode() in preview.data
    assert "预览中的裂解步骤".encode() in preview.data
    assert "核对预览和导出内容一致".encode() in preview.data
    assert b"BATCH-A-RESULT" in preview.data
    assert b"BATCH-B-RESULT" not in preview.data
    assert f'name="record_ids" value="{selected_record_id}"'.encode() in preview.data
    assert f'name="record_ids" value="{unselected_record_id}"'.encode() not in preview.data
    assert "确认导出 1 份报告".encode() in preview.data

    response = client.post("/experiment-reports/export-selected", data={
        "record_ids": [str(selected_record_id)],
        "format": "docx",
        "report_template": "notebook",
        "return_url": f"/experiment-reports?project_id={project_id}",
    })
    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        names = archive.namelist()
        report_names = [name for name in names if name.endswith(".docx")]
        assert "导出说明.txt" in names
        assert len(report_names) == 1
        assert report_names[0].startswith(
            "PROJ-REPORT-细胞应答项目/EXP-EXPORT-01-药物处理后 WB 验证/"
        )
        assert f"R-{selected_record_id}.docx" in report_names[0]
        manifest = archive.read("导出说明.txt").decode("utf-8-sig")
        assert "细胞应答项目" in manifest
        assert "报告数量：1" in manifest
        assert f"R-{selected_record_id}" in manifest
        assert f"R-{unselected_record_id}" not in manifest
        parts = _docx_parts(archive.read(report_names[0]))
        header = "".join(markup for name, markup in parts.items() if "header" in name)
        assert "LAB NOTEBOOK" in header

    legacy = client.get(f"/projects/{project_id}/reports/export?format=docx")
    assert legacy.status_code == 302
    assert legacy.headers["Location"].endswith(f"/experiment-reports?project_id={project_id}")


def test_report_preview_can_select_all_filtered_matches_and_rejects_foreign_ids(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    with app.app_context():
        experiment = db.session.get(Experiment, experiment_id)
        project = ResearchProject(
            user_id=experiment.user_id,
            title="筛选导出项目",
            code="FILTERED",
        )
        db.session.add(project)
        db.session.flush()
        experiment.project_id = project.id
        project_id = project.id
        selected_record_id = ExperimentRecord.query.filter_by(result="BATCH-A-RESULT").one().id
        db.session.commit()

    preview = client.post("/experiment-reports/export-preview", data={
        "selection_scope": "all",
        "project_id": str(project_id),
        "q": "BATCH-A-RESULT",
        "page": "1",
        "per_page": "10",
    })
    assert preview.status_code == 200
    assert "当前筛选全部结果".encode() in preview.data
    assert "已载入 1 份实验报告".encode() in preview.data
    assert b"BATCH-A-RESULT" in preview.data
    assert b"BATCH-B-RESULT" not in preview.data

    no_selection = client.post("/experiment-reports/export-preview", data={
        "selection_scope": "page",
        "project_id": str(project_id),
    })
    assert no_selection.status_code == 302
    assert no_selection.headers["Location"].endswith(
        f"/experiment-reports?project_id={project_id}"
    )

    auth.logout()
    auth.register(email="other-project-user@example.com")
    assert client.post("/experiment-reports/export-preview", data={
        "selection_scope": "page",
        "record_ids": [str(selected_record_id)],
    }).status_code == 404
    assert client.post("/experiment-reports/export-selected", data={
        "record_ids": [str(selected_record_id)],
        "format": "docx",
    }).status_code == 404
    assert client.get(f"/projects/{project_id}/reports/export?format=docx").status_code == 404


def test_experiment_file_index_collects_local_files_and_filters_metadata(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    response = client.get(f"/experiments/{experiment_id}/files?q=raw")
    assert response.status_code == 200
    assert ACTIVE_ATTACHMENT.encode() in response.data
    assert b"BATCH-A" in response.data
    assert DELETED_ATTACHMENT.encode() not in response.data


def test_json_export_groups_records_by_execution_and_keeps_flat_compatibility_view(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    response = client.get(f"/experiments/{experiment_id}/export?format=json")
    assert response.status_code == 200
    assert response.mimetype == "application/json"
    payload = json.loads(response.data)

    assert payload["schema_version"] == 3
    assert payload["experiment"]["code"] == "EXP-EXPORT-01"
    assert payload["plan_parameters"][0]["name"] == "药物浓度"
    assert payload["steps"][0]["title"] == "药物处理"

    batches = {batch["batch_code"]: batch for batch in payload["batches"]}
    assert set(batches) == {"BATCH-A", "BATCH-B"}
    assert [record["result"] for record in batches["BATCH-A"]["records"]] == ["BATCH-A-RESULT"]
    assert [record["result"] for record in batches["BATCH-B"]["records"]] == ["BATCH-B-RESULT"]

    flat_records = {record["content"]: record for record in payload["records"]}
    assert flat_records["完成第一批处理并采集原始数据。"]["batch_code"] == "BATCH-A"
    assert flat_records["完成第二批处理并采集图像。"]["batch_code"] == "BATCH-B"
    assert flat_records["完成第一批处理并采集原始数据。"]["parameters"][0]["name"] == "曝光时间"
    assert flat_records["完成第一批处理并采集原始数据。"]["attachments"][0]["relative_path"] == ACTIVE_ATTACHMENT


def test_unassigned_active_record_is_exported_once_in_explicit_history_group(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    payload = client.get(f"/experiments/{experiment_id}/export?format=json").get_json()

    assert len(payload["unassigned_records"]) == 1
    orphan = payload["unassigned_records"][0]
    assert orphan["content"] == UNASSIGNED_RECORD_CONTENT
    assert orphan["batch_id"] is None
    assert orphan["batch_code"] == "HISTORY-UNASSIGNED"
    assert sum(record["id"] == orphan["id"] for record in payload["records"]) == 1
    assert all(
        record["id"] != orphan["id"]
        for batch in payload["batches"]
        for record in batch["records"]
    )


def test_markdown_and_word_exports_show_each_execution_as_a_section(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)

    markdown = client.get(f"/experiments/{experiment_id}/export?format=markdown")
    assert markdown.status_code == 200
    report = markdown.data.decode("utf-8-sig")
    assert "## 目录" in report
    assert "## 实验概览" in report
    assert "| 字段 | 内容 |" in report
    assert "| 分类 | 文件夹 / 文件 | 版本 | 大小 | SHA-256 | 标签 | 说明 |" in report
    assert "BATCH-A" in report
    assert "BATCH-B" in report
    assert "HISTORY-UNASSIGNED" in report
    assert "过程记录" in report
    assert "人工核验" in report

    word = client.get(f"/experiments/{experiment_id}/export?format=docx")
    assert word.status_code == 200
    assert word.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    document_xml = _docx_xml(word.data)
    assert "EXP-EXPORT-01" in document_xml
    assert "BATCH-A" in document_xml
    assert "BATCH-B" in document_xml
    assert "HISTORY-UNASSIGNED" in document_xml
    assert "过程记录" in document_xml


def test_excel_export_has_execution_sheet_and_execution_code_on_record_rows(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    response = client.get(f"/experiments/{experiment_id}/export?format=xlsx")
    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    sheets = _workbook_rows(response.data)
    assert "实验批次" in sheets
    assert "过程记录" in sheets
    assert len(sheets) == 9
    assert "方案阶段" in sheets
    assert "实验步骤" in sheets
    assert "批次编号" in sheets["过程记录"][0]
    execution_code_column = sheets["过程记录"][0].index("批次编号")
    execution_codes = {row[execution_code_column] for row in sheets["过程记录"][1:]}
    assert execution_codes == {"BATCH-A", "BATCH-B", "HISTORY-UNASSIGNED"}
    assert {row[1] for row in sheets["实验批次"][1:]} == {
        "BATCH-A",
        "BATCH-B",
        "HISTORY-UNASSIGNED",
    }


def test_soft_deleted_records_and_attachments_are_excluded_from_every_format(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)

    json_response = client.get(f"/experiments/{experiment_id}/export?format=json")
    markdown_response = client.get(f"/experiments/{experiment_id}/export?format=markdown")
    word_response = client.get(f"/experiments/{experiment_id}/export?format=docx")
    excel_response = client.get(f"/experiments/{experiment_id}/export?format=xlsx")
    zip_response = client.get(f"/experiments/{experiment_id}/export?format=zip")
    assert all(response.status_code == 200 for response in (
        json_response,
        markdown_response,
        word_response,
        excel_response,
        zip_response,
    ))

    visible_text = "\n".join((
        json_response.data.decode("utf-8"),
        markdown_response.data.decode("utf-8-sig"),
        _docx_xml(word_response.data),
        _workbook_text(excel_response.data),
    ))
    assert DELETED_RECORD_CONTENT not in visible_text
    assert DELETED_ATTACHMENT not in visible_text
    assert DELETED_RECORD_ATTACHMENT not in visible_text

    with zipfile.ZipFile(io.BytesIO(zip_response.data)) as archive:
        archive_text = "\n".join((
            "\n".join(archive.namelist()),
            archive.read("report.md").decode("utf-8-sig"),
            archive.read("experiment.json").decode("utf-8"),
            archive.read("file-manifest.csv").decode("utf-8-sig"),
        ))
    assert DELETED_RECORD_CONTENT not in archive_text
    assert DELETED_ATTACHMENT not in archive_text
    assert DELETED_RECORD_ATTACHMENT not in archive_text


def test_zip_export_uses_execution_folders_and_shared_schema(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    response = client.get(f"/experiments/{experiment_id}/export?format=zip")
    assert response.status_code == 200
    assert response.mimetype == "application/zip"

    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        names = archive.namelist()
        assert "report.md" in names
        assert "experiment.json" in names
        assert "file-manifest.csv" in names
        assert any(name.startswith("files/BATCH-A/") and name.endswith(ACTIVE_ATTACHMENT) for name in names)
        assert any(name.startswith("files/BATCH-B/") and name.endswith("images/result-b.png") for name in names)
        payload = json.loads(archive.read("experiment.json"))
        manifest = archive.read("file-manifest.csv").decode("utf-8-sig")

    assert payload["schema_version"] == 3
    assert {batch["batch_code"] for batch in payload["batches"]} == {"BATCH-A", "BATCH-B"}
    assert "execution_code" in manifest.splitlines()[0]
    assert "BATCH-A" in manifest
    assert "BATCH-B" in manifest
    assert DELETED_ATTACHMENT not in manifest
    assert DELETED_RECORD_ATTACHMENT not in manifest


def test_experiment_export_rejects_unknown_format_and_other_users(client, auth, app):
    experiment_id = _create_complete_experiment(client, auth, app)
    for report_template in ("research", "notebook", "compact"):
        pdf = client.get(
            f"/experiments/{experiment_id}/export?format=pdf&report_template={report_template}"
        )
        assert pdf.status_code == 200
        assert pdf.mimetype == "application/pdf"
        assert pdf.data.startswith(b"%PDF")

    auth.logout()
    auth.register(email="other@example.com")
    for export_format in ("markdown", "pdf", "json", "docx", "xlsx", "zip"):
        assert client.get(f"/experiments/{experiment_id}/export?format={export_format}").status_code == 404
